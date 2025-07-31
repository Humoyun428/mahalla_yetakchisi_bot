import os
from openpyxl import Workbook, load_workbook

def save_report_to_excel(data, filename="data/report.xlsx"):
    if not os.path.exists(filename):
        # Fayl mavjud emas — yangi fayl yaratamiz
        wb = Workbook()
        ws = wb.active
        ws.append(["Mahalla", "Yo'nalish", "Tadbir nomi", "Vaqti", "Izoh", "Ism", "Manzil", "Rasm fayllar"])
    else:
        try:
            wb = load_workbook(filename)
            ws = wb.active
        except Exception as e:
            print(f"Xatolik: {e}")
            return

    ws = wb.active
    ws.append([
        data["mahalla"],
        data["direction"],
        data["event_name"],
        data["event_date"],
        data["description"],
        str(data["location"])  # ✅ Koordinatani string qilib yozamiz
    ])

    wb.save(filename)
