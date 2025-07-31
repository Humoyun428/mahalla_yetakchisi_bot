import os
from openpyxl import Workbook, load_workbook

def save_report_to_excel(data):
    file_path = "data/hisobotlar.xlsx"

    # Fayl mavjud bo‘lmasa — yangi yaratamiz
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Fayl IDlar", "Mahalla", "Yo‘nalish", "Vaqt", "Tadbir nomi", "Tavsif", "Latitude", "Longitude", "Qamrov"
        ])
        wb.save(file_path)

    # Yangi maʼlumot yozish
    wb = load_workbook(file_path)
    ws = wb.active
    ws.append([
        ', '.join(data["photos"]),
        data["mahalla"],
        data["direction"],
        data["date"],
        data["event_name"],
        data["description"],
        data["location"][0],
        data["location"][1],
    ])
    wb.save(file_path)